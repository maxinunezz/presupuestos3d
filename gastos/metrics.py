"""
Motor del panel de gastos.

Calcula, para un período (un mes o un año completo), el total de gastos
operativos, el desglose por categoría, la evolución mensual, el comparativo con
el período anterior, el compromiso mensual recurrente (run-rate), el resultado
operativo contra las ventas y el control de topes por categoría.

Todo el dinero se suma en Python (los totales de ventas son @property, no
columnas). A la escala del negocio es exacto y rápido.
"""

from datetime import date, datetime, time
from decimal import Decimal

from django.utils import timezone
from django.utils.translation import gettext

from .models import Gasto, TopeGasto

ZERO = Decimal("0")


def _meses():
    return [
        "",
        gettext("Enero"),
        gettext("Febrero"),
        gettext("Marzo"),
        gettext("Abril"),
        gettext("Mayo"),
        gettext("Junio"),
        gettext("Julio"),
        gettext("Agosto"),
        gettext("Septiembre"),
        gettext("Octubre"),
        gettext("Noviembre"),
        gettext("Diciembre"),
    ]


# ---------------------------------------------------------------------------
#  Rangos de fecha
# ---------------------------------------------------------------------------
def _month_range(year: int, month: int):
    start = date(year, month, 1)
    end = date(year + 1, 1, 1) if month == 12 else date(year, month + 1, 1)
    return start, end


def _year_range(year: int):
    return date(year, 1, 1), date(year + 1, 1, 1)


def _aware(d: date) -> datetime:
    return timezone.make_aware(
        datetime.combine(d, time.min), timezone.get_current_timezone()
    )


def _prev_month(year: int, month: int):
    return (year - 1, 12) if month == 1 else (year, month - 1)


# ---------------------------------------------------------------------------
#  Sumas base
# ---------------------------------------------------------------------------
def _gastos_total(start: date, end: date) -> Decimal:
    montos = Gasto.objects.filter(
        fecha__gte=start, fecha__lt=end
    ).values_list("monto", flat=True)
    return sum((Decimal(m) for m in montos), ZERO)


def _ventas_total(start: date, end: date) -> Decimal:
    """Facturación aprobada (Presupuesto.total) en el rango de fechas."""
    from budgets.models import Presupuesto

    qs = (
        Presupuesto.objects.filter(
            approved_at__gte=_aware(start), approved_at__lt=_aware(end)
        )
        .exclude(status=Presupuesto.Status.CANCELLED)
        .exclude(para_stock=True)  # reposición de stock interno, no es venta real
        .prefetch_related("items__producto")
    )
    return sum((p.total for p in qs), ZERO)


def available_years():
    """Años con datos (de gastos) + el año actual, de mayor a menor."""
    years = set(
        Gasto.objects.dates("fecha", "year").values_list("fecha__year", flat=True)
    )
    years.add(timezone.localdate().year)
    return sorted(years, reverse=True)


# ---------------------------------------------------------------------------
#  Cálculo principal
# ---------------------------------------------------------------------------
def build_gastos_metrics(year: int, month) -> dict:
    """
    `month` = 1..12 para ver un mes; None (o 0) para ver el año completo.
    Devuelve un dict con todos los números crudos (Decimals/listas).
    """
    today = timezone.localdate()
    is_month_view = bool(month)
    meses = _meses()

    if is_month_view:
        start, end = _month_range(year, month)
        months_in_period = 1
        period_label = f"{meses[month]} {year}"
        prev_y, prev_m = _prev_month(year, month)
        prev_start, prev_end = _month_range(prev_y, prev_m)
        prev_label = f"{meses[prev_m]} {prev_y}"
    else:
        start, end = _year_range(year)
        months_in_period = 12
        period_label = gettext("Año %(year)s") % {"year": year}
        prev_start, prev_end = _year_range(year - 1)
        prev_label = gettext("Año %(year)s") % {"year": year - 1}

    gastos = list(Gasto.objects.filter(fecha__gte=start, fecha__lt=end))

    # --- Total y desglose por categoría ---
    total_gastos = sum((Decimal(g.monto) for g in gastos), ZERO)
    n_gastos = len(gastos)

    por_cat = {value: {"total": ZERO, "count": 0} for value, _ in Gasto.Categoria.choices}
    for g in gastos:
        por_cat[g.categoria]["total"] += Decimal(g.monto)
        por_cat[g.categoria]["count"] += 1
    categorias = [
        {
            "value": value,
            "label": label,
            "total": por_cat[value]["total"],
            "count": por_cat[value]["count"],
            "pct": (por_cat[value]["total"] / total_gastos * 100) if total_gastos else ZERO,
        }
        for value, label in Gasto.Categoria.choices
    ]
    categorias.sort(key=lambda c: c["total"], reverse=True)

    # --- Evolución: 12 meses del año seleccionado ---
    serie = []
    for m in range(1, 13):
        ms, me = _month_range(year, m)
        tot = sum(
            (Decimal(g.monto) for g in gastos if ms <= g.fecha < me), ZERO
        )
        serie.append({"label": meses[m][:3], "month": m, "total": tot})

    # --- Comparativo con el período anterior (#3) ---
    prev_total = _gastos_total(prev_start, prev_end)
    if prev_total:
        variacion_pct = (total_gastos - prev_total) / prev_total * 100
    else:
        variacion_pct = None

    # Variación por categoría.
    prev_gastos = list(Gasto.objects.filter(fecha__gte=prev_start, fecha__lt=prev_end))
    prev_por_cat = {value: ZERO for value, _ in Gasto.Categoria.choices}
    for g in prev_gastos:
        prev_por_cat[g.categoria] += Decimal(g.monto)
    for c in categorias:
        pv = prev_por_cat.get(c["value"], ZERO)
        c["prev"] = pv
        c["var_pct"] = ((c["total"] - pv) / pv * 100) if pv else None

    # Meses con datos del período (para promediar compromiso/gasto mensual).
    # En vista de un mes es 1; en vista anual, los meses ya transcurridos del año.
    if year == today.year:
        meses_transcurridos = today.month
    elif year < today.year:
        meses_transcurridos = 12
    else:
        meses_transcurridos = 1
    run_rate_divisor = 1 if is_month_view else meses_transcurridos

    # --- Run-rate recurrente (#2) ---
    # Modelo basado en eventos: una suscripción mensual aparece como un gasto por
    # mes. El compromiso mensual = equivalente mensual sumado / meses con datos
    # (no por los 12 fijos), si no, una vista anual de año en curso lo subestima.
    recurrentes = [g for g in gastos if g.es_recurrente]
    run_rate_mensual = (
        sum((g.monthly_equivalent for g in recurrentes), ZERO) / run_rate_divisor
    ).quantize(Decimal("0.01"))
    run_rate_anual = (run_rate_mensual * 12).quantize(Decimal("0.01"))
    recurrentes_detalle = sorted(
        (
            {
                "concepto": g.concepto,
                "categoria": g.get_categoria_display(),
                "periodicidad": g.get_periodicidad_display(),
                "monthly": g.monthly_equivalent,
            }
            for g in recurrentes
        ),
        key=lambda r: r["monthly"],
        reverse=True,
    )

    # --- Resultado operativo vs ventas (#1) ---
    ventas = _ventas_total(start, end)
    resultado = ventas - total_gastos
    gastos_sobre_ventas = (total_gastos / ventas * 100) if ventas else None

    # --- Topes por categoría (#4) ---
    topes = {t.categoria: Decimal(t.monto_mensual) for t in TopeGasto.objects.all()}
    topes_rows = []
    for c in categorias:
        tope_mensual = topes.get(c["value"], ZERO)
        if tope_mensual <= 0:
            continue
        tope_periodo = tope_mensual * months_in_period
        pct = (c["total"] / tope_periodo * 100) if tope_periodo else ZERO
        topes_rows.append(
            {
                "label": c["label"],
                "gasto": c["total"],
                "tope": tope_periodo,
                "pct": pct,
                "excedido": c["total"] > tope_periodo,
                "restante": tope_periodo - c["total"],
            }
        )

    # --- Promedio mensual y acumulado anual (#6) ---
    y_start, y_end = _year_range(year)
    acumulado_anual = _gastos_total(y_start, y_end)
    promedio_mensual = (
        acumulado_anual / meses_transcurridos if meses_transcurridos else ZERO
    ).quantize(Decimal("0.01"))

    return {
        "year": year,
        "month": month or 0,
        "is_month_view": is_month_view,
        "period_label": period_label,
        "prev_label": prev_label,
        "range_str": f"{start.strftime('%d/%m/%Y')} – {(end).strftime('%d/%m/%Y')}",
        "total_gastos": total_gastos,
        "n_gastos": n_gastos,
        "categorias": categorias,
        "serie": serie,
        "prev_total": prev_total,
        "variacion_pct": variacion_pct,
        "run_rate_mensual": run_rate_mensual,
        "run_rate_anual": run_rate_anual,
        "recurrentes_detalle": recurrentes_detalle,
        "ventas": ventas,
        "resultado": resultado,
        "gastos_sobre_ventas": gastos_sobre_ventas,
        "topes_rows": topes_rows,
        "acumulado_anual": acumulado_anual,
        "promedio_mensual": promedio_mensual,
        "meses_transcurridos": meses_transcurridos,
    }


# ---------------------------------------------------------------------------
#  Formato para el template
# ---------------------------------------------------------------------------
def _money(v) -> str:
    from budgets.pdf import format_money

    return "$ " + format_money(v)


def _pct(v) -> str:
    if v is None:
        return "—"
    return f"{v:+.1f}%" if v < 0 or v > 0 else "0,0%"


def _pct_plain(v) -> str:
    if v is None:
        return "—"
    return f"{v:.1f}%"


def template_context(m: dict) -> dict:
    import json

    cat_chart = {
        "labels": [str(c["label"]) for c in m["categorias"] if c["total"] > 0],
        "data": [float(c["total"]) for c in m["categorias"] if c["total"] > 0],
    }
    serie_chart = {
        "labels": [s["label"] for s in m["serie"]],
        "data": [float(s["total"]) for s in m["serie"]],
    }

    categorias = [
        {
            "label": c["label"],
            "total": _money(c["total"]),
            "count": c["count"],
            "pct": _pct_plain(c["pct"]),
            "var_pct": _pct(c["var_pct"]),
            "var_up": (c["var_pct"] is not None and c["var_pct"] > 0),
        }
        for c in m["categorias"]
    ]
    recurrentes = [
        {
            "concepto": r["concepto"],
            "categoria": r["categoria"],
            "periodicidad": r["periodicidad"],
            "monthly": _money(r["monthly"]),
        }
        for r in m["recurrentes_detalle"]
    ]
    topes = [
        {
            "label": t["label"],
            "gasto": _money(t["gasto"]),
            "tope": _money(t["tope"]),
            "pct": _pct_plain(t["pct"]),
            "excedido": t["excedido"],
            "restante": _money(t["restante"]),
        }
        for t in m["topes_rows"]
    ]

    return {
        "period_label": m["period_label"],
        "prev_label": m["prev_label"],
        "range_str": m["range_str"],
        "total_gastos": _money(m["total_gastos"]),
        "n_gastos": m["n_gastos"],
        "categorias": categorias,
        "variacion_pct": _pct(m["variacion_pct"]),
        "variacion_up": (m["variacion_pct"] is not None and m["variacion_pct"] > 0),
        "prev_total": _money(m["prev_total"]),
        "run_rate_mensual": _money(m["run_rate_mensual"]),
        "run_rate_anual": _money(m["run_rate_anual"]),
        "recurrentes": recurrentes,
        "ventas": _money(m["ventas"]),
        "resultado": _money(m["resultado"]),
        "resultado_positivo": m["resultado"] >= 0,
        "gastos_sobre_ventas": _pct_plain(m["gastos_sobre_ventas"]),
        "topes": topes,
        "acumulado_anual": _money(m["acumulado_anual"]),
        "promedio_mensual": _money(m["promedio_mensual"]),
        "meses_transcurridos": m["meses_transcurridos"],
        "cat_chart_json": json.dumps(cat_chart),
        "serie_chart_json": json.dumps(serie_chart),
    }


# ---------------------------------------------------------------------------
#  Export a Excel (#5)
# ---------------------------------------------------------------------------
def export_xlsx(m: dict):
    """Arma el .xlsx del período de gastos y devuelve (filename, bytes)."""
    from io import BytesIO

    from openpyxl import Workbook
    from openpyxl.chart import BarChart, Reference
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    head_fill = PatternFill("solid", fgColor="111111")
    head_font = Font(color="FFFFFF", bold=True)
    title_font = Font(bold=True, size=14)

    def style_header(ws, row, ncols):
        for c in range(1, ncols + 1):
            cell = ws.cell(row=row, column=c)
            cell.fill = head_fill
            cell.font = head_font

    def autosize(ws):
        for col in ws.columns:
            width = max(
                (len(str(c.value)) for c in col if c.value is not None), default=10
            )
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(
                width + 3, 50
            )

    # --- Hoja 1: Resumen ---
    ws = wb.active
    ws.title = gettext("Resumen")
    ws["A1"] = gettext("Gastos 3darg — %(period)s (%(range)s)") % {
        "period": m["period_label"],
        "range": m["range_str"],
    }
    ws["A1"].font = title_font
    rows = [
        (gettext("Total de gastos"), float(m["total_gastos"])),
        (gettext("Cantidad de gastos"), m["n_gastos"]),
        (gettext("Total %(prev)s") % {"prev": m["prev_label"]}, float(m["prev_total"])),
        (
            gettext("Variación vs período anterior"),
            (f"{m['variacion_pct']:.1f}%" if m["variacion_pct"] is not None else "—"),
        ),
        ("", ""),
        (gettext("Compromiso mensual recurrente"), float(m["run_rate_mensual"])),
        (gettext("Proyección anual recurrente"), float(m["run_rate_anual"])),
        ("", ""),
        (gettext("Ventas del período"), float(m["ventas"])),
        (gettext("Resultado operativo (ventas − gastos)"), float(m["resultado"])),
        (
            gettext("Gastos sobre ventas"),
            (
                f"{m['gastos_sobre_ventas']:.1f}%"
                if m["gastos_sobre_ventas"] is not None
                else "—"
            ),
        ),
        ("", ""),
        (gettext("Acumulado año %(year)s") % {"year": m["year"]}, float(m["acumulado_anual"])),
        (gettext("Promedio mensual"), float(m["promedio_mensual"])),
    ]
    r = 3
    for label, value in rows:
        ws.cell(row=r, column=1, value=label)
        ws.cell(row=r, column=2, value=value)
        if value == "" and label:
            ws.cell(row=r, column=1).font = Font(bold=True)
        r += 1
    autosize(ws)

    # --- Hoja 2: Por categoría ---
    ws2 = wb.create_sheet(gettext("Por categoría"))
    ws2.append([
        gettext("Categoría"),
        gettext("Gasto"),
        gettext("% del total"),
        gettext("Cantidad"),
        m["prev_label"],
    ])
    style_header(ws2, 1, 5)
    for c in m["categorias"]:
        ws2.append(
            [
                str(c["label"]),
                float(c["total"]),
                float(c["pct"]),
                c["count"],
                float(c["prev"]),
            ]
        )
    autosize(ws2)

    # --- Hoja 3: Evolución mensual (con gráfico) ---
    ws3 = wb.create_sheet(gettext("Evolución"))
    ws3.append([gettext("Mes"), gettext("Gasto")])
    style_header(ws3, 1, 2)
    for s in m["serie"]:
        ws3.append([s["label"], float(s["total"])])
    chart = BarChart()
    chart.title = gettext("Gastos por mes — %(year)s") % {"year": m["year"]}
    chart.y_axis.title = "$"
    data = Reference(ws3, min_col=2, min_row=1, max_row=1 + len(m["serie"]))
    cats = Reference(ws3, min_col=1, min_row=2, max_row=1 + len(m["serie"]))
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.legend = None
    ws3.add_chart(chart, "D2")
    autosize(ws3)

    # --- Hoja 4: Recurrentes ---
    ws4 = wb.create_sheet(gettext("Recurrentes"))
    ws4.append([
        gettext("Concepto"),
        gettext("Categoría"),
        gettext("Periodicidad"),
        gettext("Equivalente mensual"),
    ])
    style_header(ws4, 1, 4)
    for r in m["recurrentes_detalle"]:
        ws4.append(
            [r["concepto"], str(r["categoria"]), str(r["periodicidad"]), float(r["monthly"])]
        )
    autosize(ws4)

    # --- Hoja 5: Topes ---
    ws5 = wb.create_sheet(gettext("Topes"))
    ws5.append([
        gettext("Categoría"),
        gettext("Gasto"),
        gettext("Tope"),
        gettext("% usado"),
        gettext("¿Excedido?"),
    ])
    style_header(ws5, 1, 5)
    for t in m["topes_rows"]:
        ws5.append(
            [
                str(t["label"]),
                float(t["gasto"]),
                float(t["tope"]),
                float(t["pct"]),
                gettext("Sí") if t["excedido"] else gettext("No"),
            ]
        )
    autosize(ws5)

    buffer = BytesIO()
    wb.save(buffer)
    suffix = f"{m['year']}" + (f"_{m['month']:02d}" if m["month"] else "_anual")
    fname = f"gastos_3darg_{suffix}_{timezone.localdate().isoformat()}.xlsx"
    return fname, buffer.getvalue()
