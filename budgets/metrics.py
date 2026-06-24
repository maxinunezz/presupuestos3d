"""
Motor de métricas del negocio.

Calcula KPIs de ventas, producción e inventario para un período (semana, mes o
año) y arma las series para los gráficos. Todo el dinero se suma en Python
porque `Presupuesto.total`, `PresupuestoItem.line_total` y `Producto.unit_cost`
son @property (no columnas), así que no se pueden agregar con el ORM. A la
escala del negocio esto es exacto y rápido; si algún día el histórico crece
mucho, conviene precalcular snapshots mensuales.

La fecha que define "cuándo pasó algo":
  - Ventas: approved_at (cuándo se cerró el negocio).
  - Producción: finished_at (cuándo se terminó de imprimir).
  - Compras: confirmed_at.   - Consumo: StockMovement.created_at.
"""

from collections import Counter, defaultdict
from datetime import datetime, time, timedelta
from decimal import Decimal

from django.db.models import Count
from django.utils import timezone

from .models import Presupuesto
from .pdf import format_money

# Cuántos períodos hacia atrás se grafican en la serie de facturación.
PERIODS = {
    "week": {"label": "Semana", "buckets": 8},
    "month": {"label": "Mes", "buckets": 12},
    "year": {"label": "Año", "buckets": 5},
}

ZERO = Decimal("0")


# ---------------------------------------------------------------------------
#  Períodos (buckets) en hora local
# ---------------------------------------------------------------------------
def _aware(naive: datetime) -> datetime:
    return timezone.make_aware(naive, timezone.get_current_timezone())


def period_bounds(period: str, anchor: datetime):
    """Devuelve (inicio, fin) del bucket que contiene `anchor`, en hora local."""
    local = timezone.localtime(anchor)
    d = local.date()
    if period == "week":
        monday = d - timedelta(days=local.weekday())
        start = datetime.combine(monday, time.min)
        end = start + timedelta(days=7)
    elif period == "year":
        start = datetime.combine(d.replace(month=1, day=1), time.min)
        end = datetime.combine(d.replace(year=d.year + 1, month=1, day=1), time.min)
    else:  # month
        first = d.replace(day=1)
        start = datetime.combine(first, time.min)
        nxt = (
            first.replace(year=first.year + 1, month=1)
            if first.month == 12
            else first.replace(month=first.month + 1)
        )
        end = datetime.combine(nxt, time.min)
    return _aware(start), _aware(end)


def _bucket_label(period: str, start: datetime) -> str:
    local = timezone.localtime(start)
    if period == "week":
        return "Sem " + local.strftime("%d/%m")
    if period == "year":
        return local.strftime("%Y")
    return local.strftime("%m/%Y")


def make_buckets(period: str, now: datetime):
    """Lista de buckets [más viejo ... actual] para la serie del período."""
    n = PERIODS[period]["buckets"]
    start, end = period_bounds(period, now)
    out = []
    for _ in range(n):
        out.append({"start": start, "end": end, "label": _bucket_label(period, start)})
        end = start
        start = period_bounds(period, start - timedelta(seconds=1))[0]
    out.reverse()
    return out


# ---------------------------------------------------------------------------
#  Cálculo de métricas
# ---------------------------------------------------------------------------
def _approved_in_range(start, end):
    """Presupuestos aprobados en [start, end) con todo prefetch para no hacer N+1."""
    return list(
        Presupuesto.objects.filter(
            approved_at__gte=start, approved_at__lt=end
        ).prefetch_related(
            "items__producto__filament_lines__filament",
            "items__producto__aggregate_lines__aggregate",
        )
    )


def build_metrics(period: str, now: datetime = None) -> dict:
    """
    Devuelve un dict con todas las métricas crudas (Decimals/listas) del período
    actual + la serie de facturación. La vista lo formatea para el template y el
    Excel.
    """
    if period not in PERIODS:
        period = "month"
    now = now or timezone.now()

    buckets = make_buckets(period, now)
    cur = buckets[-1]
    range_start = buckets[0]["start"]
    range_end = buckets[-1]["end"]

    approved_all = _approved_in_range(range_start, range_end)
    cur_approved = [p for p in approved_all if cur["start"] <= p.approved_at < cur["end"]]

    # --- Serie de facturación por bucket (para el gráfico) ---
    serie = []
    for b in buckets:
        total = sum(
            (p.total for p in approved_all if b["start"] <= p.approved_at < b["end"]),
            ZERO,
        )
        serie.append({"label": b["label"], "total": total})

    # =====================  A) VENTAS  =====================
    facturacion = sum((p.total for p in cur_approved), ZERO)
    n_aprobados = len(cur_approved)
    ticket = (facturacion / n_aprobados) if n_aprobados else ZERO

    # Conversión por cohorte: de los enviados en el período, cuántos se aprobaron.
    sent_qs = Presupuesto.objects.filter(
        sent_at__gte=cur["start"], sent_at__lt=cur["end"]
    )
    n_enviados = sent_qs.count()
    n_conv = sent_qs.filter(approved_at__isnull=False).count()
    conversion = (n_conv / n_enviados * 100) if n_enviados else None

    # Embudo: foto global del pipeline por estado actual.
    status_counts = dict(
        Presupuesto.objects.values_list("status").annotate(c=Count("id"))
    )
    embudo = [
        {"label": label, "count": status_counts.get(value, 0)}
        for value, label in Presupuesto.Status.choices
    ]

    # Ranking de productos (período actual): por cantidad y por $.
    by_qty = Counter()
    by_money = defaultdict(lambda: ZERO)
    cli_money = defaultdict(lambda: ZERO)
    revenue = ZERO
    cost = ZERO
    for p in cur_approved:
        cli_money[p.client_name] += p.total
        for it in p.items.all():
            name = it.producto.name
            by_qty[name] += it.quantity
            line = it.line_total
            by_money[name] += line
            revenue += line
            cost += Decimal(it.quantity) * it.producto.unit_cost

    top_productos_qty = by_qty.most_common(10)
    top_productos_money = sorted(by_money.items(), key=lambda kv: kv[1], reverse=True)[:10]
    top_clientes = sorted(cli_money.items(), key=lambda kv: kv[1], reverse=True)[:10]

    margen_pct = ((revenue - cost) / revenue * 100) if revenue else None

    # Tiempo de ciclo: días approved_at -> completed_at (completados en el período).
    completed = Presupuesto.objects.filter(
        completed_at__gte=cur["start"],
        completed_at__lt=cur["end"],
        approved_at__isnull=False,
    )
    ciclo_days = [
        (p.completed_at - p.approved_at).total_seconds() / 86400 for p in completed
    ]
    tiempo_ciclo = (sum(ciclo_days) / len(ciclo_days)) if ciclo_days else None

    # =====================  B) PRODUCCIÓN  =====================
    from production.models import ProductionJob

    done = list(
        ProductionJob.objects.filter(
            status=ProductionJob.Status.DONE,
            finished_at__gte=cur["start"],
            finished_at__lt=cur["end"],
        ).select_related("producto", "machine")
    )
    piezas_impresas = sum(j.quantity for j in done)
    horas_impresas = sum((j.print_hours for j in done), ZERO)

    maq = defaultdict(lambda: {"jobs": 0, "horas": ZERO, "piezas": 0})
    for j in done:
        key = j.machine.name if j.machine else "Sin máquina"
        maq[key]["jobs"] += 1
        maq[key]["horas"] += j.print_hours
        maq[key]["piezas"] += j.quantity
    uso_maquinas = sorted(
        ({"name": k, **v} for k, v in maq.items()),
        key=lambda r: r["horas"],
        reverse=True,
    )

    from inventory.models import StockMovement

    reprints = StockMovement.objects.filter(
        reason=StockMovement.Reason.REPRINT_FAILURE,
        created_at__gte=cur["start"],
        created_at__lt=cur["end"],
    ).count()
    n_done = len(done)
    reprint_rate = (reprints / n_done * 100) if n_done else None

    # Cumplimiento de entrega: completados en el período con due_date.
    ent = Presupuesto.objects.filter(
        completed_at__gte=cur["start"],
        completed_at__lt=cur["end"],
        due_date__isnull=False,
    )
    total_ent = ent.count()
    on_time = sum(1 for p in ent if p.completed_at <= p.due_date)
    cumplimiento = (on_time / total_ent * 100) if total_ent else None

    # =====================  C) INVENTARIO / COSTOS  =====================
    from inventory.models import Aggregate, Compra, Filament

    compras = (
        Compra.objects.filter(
            status=Compra.Status.CONFIRMED,
            confirmed_at__gte=cur["start"],
            confirmed_at__lt=cur["end"],
        ).prefetch_related("lines__filament", "lines__aggregate")
    )
    gasto_compras = sum((c.total for c in compras), ZERO)
    n_compras = compras.count()

    prod_mov = StockMovement.objects.filter(
        reason=StockMovement.Reason.PRODUCTION,
        created_at__gte=cur["start"],
        created_at__lt=cur["end"],
    ).select_related("filament", "aggregate")
    fil_grams = ZERO
    consumo_valor = ZERO
    for m in prod_mov:
        q = abs(m.quantity)
        if m.filament_id:
            fil_grams += q
            consumo_valor += q * m.filament.cost_per_gram
        elif m.aggregate_id:
            consumo_valor += q * m.aggregate.cost_per_unit

    low_stock = sum(1 for f in Filament.objects.filter(is_active=True) if f.is_low_stock)
    low_stock += sum(1 for a in Aggregate.objects.filter(is_active=True) if a.is_low_stock)

    return {
        "period": period,
        "period_label": PERIODS[period]["label"],
        "cur_start": cur["start"],
        "cur_end": cur["end"],
        "now": now,
        "serie": serie,
        # Ventas
        "facturacion": facturacion,
        "n_aprobados": n_aprobados,
        "ticket": ticket,
        "n_enviados": n_enviados,
        "n_conv": n_conv,
        "conversion": conversion,
        "embudo": embudo,
        "top_productos_qty": top_productos_qty,
        "top_productos_money": top_productos_money,
        "top_clientes": top_clientes,
        "tiempo_ciclo": tiempo_ciclo,
        # Producción
        "piezas_impresas": piezas_impresas,
        "horas_impresas": horas_impresas,
        "uso_maquinas": uso_maquinas,
        "reprints": reprints,
        "reprint_rate": reprint_rate,
        "cumplimiento": cumplimiento,
        "total_ent": total_ent,
        # Inventario / costos
        "gasto_compras": gasto_compras,
        "n_compras": n_compras,
        "fil_grams": fil_grams,
        "consumo_valor": consumo_valor,
        "margen_pct": margen_pct,
        "low_stock": low_stock,
    }


# ---------------------------------------------------------------------------
#  Helpers de formato
# ---------------------------------------------------------------------------
def _money(value) -> str:
    return "$ " + format_money(value or 0)


def _pct(value) -> str:
    return "—" if value is None else f"{value:.1f}%"


def _hours(value) -> str:
    return f"{Decimal(value or 0):.1f} h"


def _days(value) -> str:
    return "—" if value is None else f"{value:.1f} días"


def _period_range_str(m) -> str:
    a = timezone.localtime(m["cur_start"]).strftime("%d/%m/%Y")
    b = timezone.localtime(m["cur_end"] - timedelta(seconds=1)).strftime("%d/%m/%Y")
    return f"{a} – {b}"


# ---------------------------------------------------------------------------
#  Export a Excel (openpyxl)
# ---------------------------------------------------------------------------
def export_xlsx(m: dict):
    """Arma el .xlsx del período y devuelve (filename, bytes)."""
    from io import BytesIO

    from openpyxl import Workbook
    from openpyxl.chart import BarChart, Reference
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    head_fill = PatternFill("solid", fgColor="111111")
    head_font = Font(color="FFFFFF", bold=True)
    title_font = Font(bold=True, size=14)

    def style_header(ws, row, ncols):
        for c in range(1, ncols + 1):
            cell = ws.cell(row=row, column=c)
            cell.fill = head_fill
            cell.font = head_font

    def autosize(ws):
        for col in ws.columns:
            width = max((len(str(c.value)) for c in col if c.value is not None), default=10)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(width + 3, 50)

    # --- Hoja 1: Resumen ---
    ws = wb.active
    ws.title = "Resumen"
    ws["A1"] = f"Métricas 3darg — {m['period_label']} ({_period_range_str(m)})"
    ws["A1"].font = title_font
    rows = [
        ("VENTAS", ""),
        ("Facturación aprobada", _money(m["facturacion"])),
        ("Presupuestos aprobados", m["n_aprobados"]),
        ("Ticket promedio", _money(m["ticket"])),
        ("Enviados / Aprobados", f'{m["n_enviados"]} / {m["n_conv"]}'),
        ("Tasa de conversión", _pct(m["conversion"])),
        ("Margen bruto", _pct(m["margen_pct"])),
        ("Tiempo de ciclo (aprob.→entrega)", _days(m["tiempo_ciclo"])),
        ("", ""),
        ("PRODUCCIÓN", ""),
        ("Piezas impresas", m["piezas_impresas"]),
        ("Horas impresas", _hours(m["horas_impresas"])),
        ("Reimpresiones por falla", m["reprints"]),
        ("Tasa de reimpresión", _pct(m["reprint_rate"])),
        ("Cumplimiento de entrega", _pct(m["cumplimiento"])),
        ("", ""),
        ("INVENTARIO / COSTOS", ""),
        ("Gasto en compras", _money(m["gasto_compras"])),
        ("N° de compras", m["n_compras"]),
        ("Consumo de material ($)", _money(m["consumo_valor"])),
        ("Filamento consumido (g)", f'{m["fil_grams"]:.0f}'),
        ("Insumos bajo stock mínimo", m["low_stock"]),
    ]
    r = 3
    for label, value in rows:
        ws.cell(row=r, column=1, value=label)
        ws.cell(row=r, column=2, value=value)
        if value == "" and label:
            ws.cell(row=r, column=1).font = Font(bold=True)
        r += 1
    autosize(ws)

    # --- Hoja 2: Facturación por período (con gráfico nativo) ---
    ws2 = wb.create_sheet("Facturación")
    ws2.append([m["period_label"], "Facturación"])
    style_header(ws2, 1, 2)
    for s in m["serie"]:
        ws2.append([s["label"], float(s["total"])])
    if len(m["serie"]) > 1:
        chart = BarChart()
        chart.title = "Facturación por " + m["period_label"].lower()
        chart.y_axis.title = "$"
        data = Reference(ws2, min_col=2, min_row=1, max_row=1 + len(m["serie"]))
        cats = Reference(ws2, min_col=1, min_row=2, max_row=1 + len(m["serie"]))
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.legend = None
        ws2.add_chart(chart, "D2")
    autosize(ws2)

    # --- Hoja 3: Ranking de productos ---
    ws3 = wb.create_sheet("Productos")
    ws3.append(["Producto", "Cantidad vendida"])
    style_header(ws3, 1, 2)
    for name, qty in m["top_productos_qty"]:
        ws3.append([name, qty])
    ws3.append([])
    base = ws3.max_row + 1
    ws3.append(["Producto", "Facturación"])
    style_header(ws3, base, 2)
    for name, money in m["top_productos_money"]:
        ws3.append([name, float(money)])
    autosize(ws3)

    # --- Hoja 4: Top clientes ---
    ws4 = wb.create_sheet("Clientes")
    ws4.append(["Cliente", "Facturación"])
    style_header(ws4, 1, 2)
    for name, money in m["top_clientes"]:
        ws4.append([name, float(money)])
    autosize(ws4)

    # --- Hoja 5: Producción por máquina ---
    ws5 = wb.create_sheet("Producción")
    ws5.append(["Máquina", "Trabajos", "Piezas", "Horas impresas"])
    style_header(ws5, 1, 4)
    for row in m["uso_maquinas"]:
        ws5.append([row["name"], row["jobs"], row["piezas"], float(row["horas"])])
    autosize(ws5)

    buffer = BytesIO()
    wb.save(buffer)
    fname = f'metricas_3darg_{m["period"]}_{timezone.localdate().isoformat()}.xlsx'
    return fname, buffer.getvalue()


# ---------------------------------------------------------------------------
#  Contexto para el template (incluye datos de los gráficos en JSON)
# ---------------------------------------------------------------------------
def template_context(m: dict) -> dict:
    import json

    fact_chart = {
        "labels": [s["label"] for s in m["serie"]],
        "data": [float(s["total"]) for s in m["serie"]],
    }
    embudo_chart = {
        "labels": [e["label"] for e in m["embudo"]],
        "data": [e["count"] for e in m["embudo"]],
    }
    maq_chart = {
        "labels": [r["name"] for r in m["uso_maquinas"]],
        "data": [float(r["horas"]) for r in m["uso_maquinas"]],
    }

    return {
        "period": m["period"],
        "period_label": m["period_label"],
        "range_str": _period_range_str(m),
        # KPIs ventas
        "facturacion": _money(m["facturacion"]),
        "n_aprobados": m["n_aprobados"],
        "ticket": _money(m["ticket"]),
        "conversion": _pct(m["conversion"]),
        "conversion_detail": f'{m["n_conv"]}/{m["n_enviados"]} enviados',
        "margen_pct": _pct(m["margen_pct"]),
        "tiempo_ciclo": _days(m["tiempo_ciclo"]),
        # KPIs producción
        "piezas_impresas": m["piezas_impresas"],
        "horas_impresas": _hours(m["horas_impresas"]),
        "reprint_rate": _pct(m["reprint_rate"]),
        "reprints": m["reprints"],
        "cumplimiento": _pct(m["cumplimiento"]),
        "total_ent": m["total_ent"],
        # KPIs inventario / costos
        "gasto_compras": _money(m["gasto_compras"]),
        "n_compras": m["n_compras"],
        "consumo_valor": _money(m["consumo_valor"]),
        "fil_grams": f'{m["fil_grams"]:.0f} g',
        "low_stock": m["low_stock"],
        # Tablas
        "embudo": m["embudo"],
        "top_productos_qty": [{"name": n, "qty": q} for n, q in m["top_productos_qty"]],
        "top_productos_money": [
            {"name": n, "money": _money(v)} for n, v in m["top_productos_money"]
        ],
        "top_clientes": [{"name": n, "money": _money(v)} for n, v in m["top_clientes"]],
        "uso_maquinas": [
            {
                "name": r["name"],
                "jobs": r["jobs"],
                "piezas": r["piezas"],
                "horas": _hours(r["horas"]),
            }
            for r in m["uso_maquinas"]
        ],
        # Gráficos (JSON)
        "fact_chart_json": json.dumps(fact_chart),
        "embudo_chart_json": json.dumps(embudo_chart),
        "maq_chart_json": json.dumps(maq_chart),
    }
